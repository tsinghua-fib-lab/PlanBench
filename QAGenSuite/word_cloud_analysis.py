import json
from wordcloud import WordCloud
import jieba
import pandas as pd
import numpy as np
import wordcloud
from PIL import Image
import matplotlib.pyplot as plt
from openpyxl import load_workbook

def transform(file):
    df = pd.read_excel(file)
    df = df[['题号', '问题', 'A', 'B', 'C', 'D', 'E', '答案', '解析']]
    df.columns = ['id', 'question', 'A', 'B', 'C', 'D', 'E', 'answer', 'explanation']
    df['id'] = df['id'] - 1
    if 'basic' in file:
        df['type'] = 'basic'
    elif 'knowledge' in file:
        df['type'] = 'knowledge'
    else:
        df['type'] = 'regulation'
    return df


def generate_sft_data(df,results_single,results_mul):
    correct_message = []
    error_message = []
    for i, row in df.iterrows():
        if row['type'] == 'basic':
            if row['id'] < 80:
                instruction = f"{row['question']}A. {row['A']}B. {row['B']}C. {row['C']}D. {row['D']}"
                if results_single['urban_and_rural_planner_basic'][str(row['id'])]==row['answer']:
                    correct_message.append(instruction)
                else:
                    error_message.append(instruction)
            else:
                instruction = f"{row['question']}A. {row['A']}B. {row['B']}C. {row['C']}D. {row['D']}E. {row['E']}"
                if results_mul['urban_and_rural_planner_basic_multi'][str(row['id']-80)]==row['answer']:
                    correct_message.append(instruction)
                else:
                    error_message.append(instruction)
        elif row['type'] == 'knowledge':
            if row['id'] < 80:
                instruction = f"{row['question']}A. {row['A']}B. {row['B']}C. {row['C']}D. {row['D']}"
                if results_single['urban_and_rural_planner_knowledge'][str(row['id'])] == row['answer']:
                    correct_message.append(instruction)
                else:
                    error_message.append(instruction)
            else:
                instruction = f"{row['question']}A. {row['A']}B. {row['B']}C. {row['C']}D. {row['D']}E. {row['E']}"
                if results_mul['urban_and_rural_planner_knowledge_multi'][str(row['id'] - 80)] == row['answer']:
                    correct_message.append(instruction)
                else:
                    error_message.append(instruction)
        elif row['type'] == 'regulation':
            if row['id'] < 80:
                instruction = f"{row['question']}A. {row['A']}B. {row['B']}C. {row['C']}D. {row['D']}"
                if results_single['urban_and_rural_planner_regulation'][str(row['id'])] == row['answer']:
                    correct_message.append(instruction)
                else:
                    error_message.append(instruction)
            else:
                instruction = f"{row['question']}A. {row['A']}B. {row['B']}C. {row['C']}D. {row['D']}E. {row['E']}"
                if results_mul['urban_and_rural_planner_regulation_multi'][str(row['id'] - 80)] == row['answer']:
                    correct_message.append(instruction)
                else:
                    error_message.append(instruction)
    return correct_message,error_message


if __name__ == '__main__':
    data_dir = 'D:\\python\\PDFprocess\\'

    with open('results_single.json', 'r', encoding='utf-8') as f:
        results_single = json.load(f)
    with open('results_mul.json', 'r', encoding='utf-8') as f:
        results_mul = json.load(f)

    files = ['D:\\python\\PDFprocess\\urban_and_rural_planner_basic_test.xlsx','urban_and_rural_planner_knowledge_test.xlsx','urban_and_rural_planner_regulation_test.xlsx']
    correct_data = []
    error_data = []
    for file in files:
        df = transform(file)
        correct_message,error_message = generate_sft_data(df,results_single,results_mul)
        correct_data.extend(correct_message)
        error_data.extend(error_message)
correct = ' '.join(correct_data)
error = ' '.join(error_data)

jieba.load_userdict('D:\\python\\PDFprocess\\dict.txt')
words_correct = jieba.lcut(correct,cut_all=False )
words_error = jieba.lcut(error,cut_all=False)

words_correct = [word for word in words_correct if len(word) != 2 or len(word) == 1 ]
words_error = [word for word in words_error if len(word) != 2 or len(word) == 1 ]

wordsDictCorrect = {}
wordsDictError = {}
for word in words_correct:
    if len(word) == 1:
        continue
    else:
        wordsDictCorrect.setdefault(word, 0)
        wordsDictCorrect[word] += 1

for word in words_error:
    if len(word) == 1:
        continue
    else:
        wordsDictError.setdefault(word, 0)
        wordsDictError[word] += 1

stopWords = ["______", "关于", "下列", "根据", "错误", "正确", "说法", "属于", "表述", "实施",
             "要求", "准确", "以上", "处理", "通过","国家","城市","规划","空间","用地","建设",
             "保护","设施","设置","交通","应当","土地","低于","有关","可以","使用"]
for word in stopWords:
    if word in wordsDictCorrect:
        del wordsDictCorrect[word]
    if word in wordsDictError:
        del wordsDictError[word]



wordsDictCorrect_seq = sorted(wordsDictCorrect.items(), key=lambda x: x[1], reverse=True)
wordsDictError_seq = sorted(wordsDictError.items(), key=lambda x: x[1], reverse=True)
print(wordsDictCorrect_seq)
print(wordsDictError_seq)
dfCorrect = pd.DataFrame(wordsDictCorrect_seq,columns=['词','次数'])
dfCorrect.to_excel("正确词频.xlsx",index = False)
dfError = pd.DataFrame(wordsDictError_seq,columns=['词','次数'])
dfError.to_excel("错误词频.xlsx",index = False)

files = ['D:\python\PDFprocess\原始词频\\正确词频（译文）.xlsx','D:\python\PDFprocess\原始词频\\错误词频（译文）.xlsx',
         'D:\python\PDFprocess\去除二字词后的词频\\正确词频（译文）.xlsx','D:\python\PDFprocess\去除二字词后的词频\\错误词频（译文）.xlsx']

for file in files:
    wb = load_workbook(file)
    ws = wb.active
    wordFreq = {}
    for i in range(2, ws.max_row + 1):
        word = ws["A" + str(i)].value
        freq = ws["B" + str(i)].value
        wordFreq[word] = freq

    wc = WordCloud(font_path='simhei.ttf',
                   background_color='white',
                   max_words=1000,
                   width=1000,
                   height=600)


    wc.generate_from_frequencies(wordFreq)

    plt.imshow(wc, interpolation='bilinear')
    plt.axis('off')
    plt.show()

